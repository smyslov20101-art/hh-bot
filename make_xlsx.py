"""
Сгенерировать ВАКАНСИИ.xlsx с цветами и выпадающими списками статусов.

Цвета:
- Новая → белый
- Откликнулся → жёлтый
- Прошёл анкету / собес / интервью → зелёный
- Отказали → красный
"""

import sys
sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule


# Статусы для выпадающего списка
STATUSES = [
    "Новая",
    "Откликнулся",
    "Заполнил анкету",
    "Прошёл собес",
    "Прошёл AI-интервью",
    "Ждём ответа",
    "Тестовое задание",
    "Отказали",
    "Принят",
]

# Цвета (ARGB hex)
COLOR_NEW = "FFFFFF"          # белый
COLOR_APPLIED = "FFF2CC"      # жёлтый
COLOR_IN_PROGRESS = "D9EAD3"  # зелёный
COLOR_REJECTED = "F4CCCC"     # красный

# Данные вакансий
DATA = [
    # (Tier, Компания, Вакансия, Зарплата, Статус, Ссылка, Заметки, Письмо)
    ("1", "Манн-Рудницкая", "AI Engineer / Специалист по AI-агентам", "120-150к",
     "Откликнулся", "https://hh.ru/vacancy/133120704",
     "Claude Code — основной инструмент. Гибрид Москва.", ""),

    ("1", "Qseller (ИП Мищенко)", "Вайб-кодер на Claude Code", "от 150к",
     "Откликнулся", "https://hh.ru/vacancy/132807106",
     "Claude Code в требованиях. Полная удалёнка.", ""),

    ("1", "IQBIQ", "AI Automation Builder", "120-250к",
     "Откликнулся", "https://hh.ru/vacancy/132836189",
     "AI автоматизация. Удалёнка.", ""),

    ("1", "БестВей", "AI-специалист / Prompt-инженер", "от 100к",
     "Откликнулся", "https://hh.ru/vacancy/133450899",
     "Python + API = бонус. Удалёнка.", ""),

    ("1", "Безлимит", "AI Transformation Lead / AI Evangelist", "от 200к",
     "Ждём ответа", "https://hh.ru/vacancy/133455243",
     "Claude в стеке. Опыт не требуется. Офис в МСК.", ""),

    ("1", "Иванова Марина (Большая Миссия)", "AI Product & Operations Lead", "от 120к",
     "Откликнулся", "https://hh.ru/vacancy/133486511",
     "SKOLKOVO 100. Claude. Удалёнка.", ""),

    ("1", "Galera Club", "AI-инженер / Вайбкодер / ИИ программист", "от 60к (3ч/день)",
     "Откликнулся", "https://hh.ru/vacancy/133482220",
     "Claude Code в требованиях. Парт-тайм 3ч/день.", ""),

    ("1", "ООО РЕКРИО", "Разработчик AI-интеграций (n8n / Cloude Code)", "60к + 13% (до 250к)",
     "Новая", "https://hh.ru/vacancy/133085355",
     "Cloude Code прямо в названии. AI-интеграции + Python. Офис МСК.",
     "Здравствуйте!\n\nРаботаю с Claude Code как основным инструментом ежедневно.\n\n"
     "Портфолио:\n\n"
     "idea-to-code-bot — AI-агент: идея текстом → план → ZIP с рабочим кодом\n"
     "(Python, aiogram 3, Claude API, async)\n"
     "github.com/smyslov20101-art/idea-to-code-bot\n\n"
     "wife-bot — Telegram-бот с LLM + автоматические уведомления\n"
     "(Python, SQLite, python-telegram-bot, VPS)\n"
     "github.com/smyslov20101-art/wife-bot\n\n"
     "Стек: Python, Claude Code, REST API, JSON, webhooks, базовый SQL,\n"
     "Git, VS Code, деплой на VPS. Интересна именно AI-разработка и интеграции.\n"
     "Готов к тестовому заданию.\n\n"
     "GitHub: github.com/smyslov20101-art"),

    ("2", "Бизнес Решение", "Python-разработчик / AI", "не указана",
     "Прошёл AI-интервью", "",
     "Прошёл AI-интервью. Ждём результата.", ""),

    ("2", "Где радость", "Специалист по AI-автоматизациям", "не указана",
     "Заполнил анкету", "",
     "Заполнил 3-страничную анкету. Ждём.", ""),

    ("2", "Sebekon", "AI-интегратор / AI Automation Engineer", "не указана",
     "Отказали", "https://hh.ru/vacancy/133303355",
     "Отказали — не критично.", ""),

    ("2", "ИП Кириченко", "AI-специалист в онлайн-школу", "от 85к",
     "Откликнулся", "https://hh.ru/vacancy/133464908",
     "Claude в стеке. 6 часов в день. Удалёнка.", ""),

    ("2", "Стрит Лайт", "Инженер внедрения", "от 70к",
     "Откликнулся", "https://hh.ru/vacancy/132814689",
     "Claude Code в требованиях. ГПХ.", ""),

    ("2", "Renera Development", "Специалист по внедрению AI-инструментов", "250000",
     "Новая", "https://hh.ru/vacancy/133058376",
     "250к. Claude Code в стеке. ОФИС МСК. Опыт 3-6 лет — выше твоего но требования мягкие.",
     "Здравствуйте!\n\nРаботаю с Claude Code, ChatGPT и AI-инструментами ежедневно —\n"
     "использую их для разработки реальных AI-продуктов и автоматизации.\n\n"
     "Портфолио:\n\n"
     "idea-to-code-bot — AI-агент на Claude API: автоматизация генерации\n"
     "кода по текстовому описанию (Python, aiogram 3, Claude API).\n"
     "github.com/smyslov20101-art/idea-to-code-bot\n\n"
     "wife-bot — Telegram-бот с LLM + автоуведомления (Python, SQLite, VPS).\n"
     "github.com/smyslov20101-art/wife-bot\n\n"
     "Понимаю как применять AI в бизнес-процессах: Second Brain,\n"
     "автоматизация протоколов, отчётов, интеграции с Google Sheets/Telegram/CRM.\n"
     "Умею объяснять технологии простым языком — могу обучать сотрудников.\n"
     "Готов к тестовому заданию.\n\n"
     "GitHub: github.com/smyslov20101-art"),

    ("3", "ИП Аксенов", "Специалист по нейросетям (репетитор)", "65-118к",
     "Откликнулся", "https://hh.ru/vacancy/133436384",
     "Репетитор. 2-3 часа/день. 1500р/ч.", ""),

    ("3", "AGIMA", "Junior Python-разработчик", "не указана",
     "Ждём ответа", "",
     "Рекрутинговое агентство. Диалог с ботом.", ""),

    ("3", "Green Code", "Python-разработчик", "не указана",
     "Откликнулся", "", "Ждём ответа.", ""),

    ("3", "hitalent", "Junior разработчик", "не указана",
     "Откликнулся", "", "Ждём ответа.", ""),

    ("3", "ООО Поколение", "Python-разработчик", "не указана",
     "Откликнулся", "", "Ждём ответа.", ""),

    ("3", "ХАБ (hub-app.ru)", "Специалист No-code / AI", "—",
     "Заполнил анкету", "https://hub-app.ru",
     "Биржа специалистов. Заполнил профиль.", ""),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"

    # ── Заголовки ─────────────────────────────────────────
    headers = ["№", "Tier", "Компания", "Вакансия", "Зарплата",
               "Статус", "Ссылка", "Заметки", "Письмо"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Стиль для кликабельной ссылки
    link_font = Font(color="0563C1", underline="single")

    # ── Данные ────────────────────────────────────────────
    for row_idx, row in enumerate(DATA, start=2):
        tier, company, name, salary, status, url, notes, letter = row
        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # №
        ws.cell(row=row_idx, column=2, value=tier)
        ws.cell(row=row_idx, column=3, value=company)
        ws.cell(row=row_idx, column=4, value=name)
        ws.cell(row=row_idx, column=5, value=salary)
        ws.cell(row=row_idx, column=6, value=status)

        # Ссылка кликабельная: HYPERLINK формула + стиль
        link_cell = ws.cell(row=row_idx, column=7)
        if url:
            link_cell.value = f'=HYPERLINK("{url}","{url}")'
            link_cell.font = link_font
        ws.cell(row=row_idx, column=8, value=notes)
        ws.cell(row=row_idx, column=9, value=letter)

        # перенос текста на длинных полях
        ws.cell(row=row_idx, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Выпадающий список на колонке Статус (F) ───────────
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUSES) + '"',
        allow_blank=True,
    )
    dv.error = "Выберите статус из списка"
    dv.errorTitle = "Недопустимое значение"
    dv.prompt = "Выберите статус"
    dv.promptTitle = "Статус"
    ws.add_data_validation(dv)
    dv.add(f"F2:F{len(DATA) + 1}")

    # ── Условное форматирование по цветам ─────────────────
    yellow_fill = PatternFill("solid", fgColor=COLOR_APPLIED)
    green_fill = PatternFill("solid", fgColor=COLOR_IN_PROGRESS)
    red_fill = PatternFill("solid", fgColor=COLOR_REJECTED)

    # Применяем ко всей строке (A:I) на основе значения в F
    last_row = len(DATA) + 1
    rng = f"A2:I{last_row}"

    # Жёлтый — Откликнулся
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=['$F2="Откликнулся"'], fill=yellow_fill),
    )
    # Зелёный — анкета/собес/интервью/ждём
    for status_value in ("Заполнил анкету", "Прошёл собес", "Прошёл AI-интервью",
                         "Ждём ответа", "Тестовое задание", "Принят"):
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'$F2="{status_value}"'], fill=green_fill),
        )
    # Красный — отказали
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=['$F2="Отказали"'], fill=red_fill),
    )

    # ── Ширина колонок ────────────────────────────────────
    widths = {
        "A": 5,    # №
        "B": 8,    # Tier
        "C": 28,   # Компания
        "D": 40,   # Вакансия
        "E": 18,   # Зарплата
        "F": 22,   # Статус
        "G": 38,   # Ссылка
        "H": 50,   # Заметки
        "I": 70,   # Письмо
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # высота строк побольше для писем
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = 80

    # Заморозить первую строку
    ws.freeze_panes = "A2"

    # Тонкие границы
    thin = Side(border_style="thin", color="CCCCCC")
    for row in ws.iter_rows(min_row=1, max_row=last_row, max_col=9):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Сохранение ────────────────────────────────────────
    out_path = Path("C:/Users/mim/Desktop/ВАКАНСИИ.xlsx")
    wb.save(out_path)
    print(f"✅ Сохранено: {out_path}")
    print(f"   Записей: {len(DATA)}")


if __name__ == "__main__":
    main()
